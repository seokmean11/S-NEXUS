"""Mark ERP Excel cells that need review (invalid codes + team mismatches)."""
from __future__ import annotations

import shutil
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "tmp" / "2026.07.31.xlsx"
SRC_NET = Path(
    r"\\krslshdb06\03_경영기획본부\원가관리팀\원가정보팀전용"
    r"\공통데이타\PPM(DB)\프로젝트등록파일(ERP)\2026.07.31.xlsx"
)
OUT = ROOT / "2026.07.31_검토필요.xlsx"

INVALID_CODES = {
    "2025-3005-42": "코드 형식 오류: 마지막 단계 자리(4)는 웹에서 공모/설계/제작(1~3)만 허용",
    "2026-8901-30": "코드 형식 오류: 마지막 단계 자리(4)는 웹에서 공모/설계/제작(1~3)만 허용",
    "2028-3005-43": "코드 형식 오류: 마지막 단계 자리(4)는 웹에서 공모/설계/제작(1~3)만 허용",
}

TEAM_MISMATCH = {
    "2024-3012-10": '담당팀 "문화기술연구소" → 코드 분류(해외) 팀 목록과 불일치',
    "2025-1024-10": '담당팀 "스튜디오스페이스타임" → 코드 분류(전시) 팀 목록과 불일치',
    "2025-1022-10": '담당팀 "스튜디오스페이스타임" → 코드 분류(전시) 팀 목록과 불일치',
    "2025-1022-30": '담당팀 "스튜디오스페이스타임" → 코드 분류(전시) 팀 목록과 불일치',
    "2024-3012-30": '담당팀 "해외사업실" → 코드 분류(해외) 팀 목록과 불일치',
    "2025-1060-20": '담당팀 "스튜디오스페이스타임" → 코드 분류(전시) 팀 목록과 불일치',
}

FILL_CODE = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FILL_TEAM = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
FILL_HEADER = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")


def ensure_source() -> Path:
    if SRC.exists():
        return SRC
    SRC.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(SRC_NET, SRC)
    return SRC


def main() -> None:
    wb = openpyxl.load_workbook(ensure_source())
    ws = wb.active

    review_col = ws.max_column + 1
    ws.cell(row=1, column=review_col, value="검토사유")
    ws.cell(row=1, column=review_col).fill = FILL_HEADER
    ws.cell(row=1, column=review_col).font = Font(bold=True)

    marked_code = 0
    marked_team = 0

    for row_idx in range(2, ws.max_row + 1):
        code_value = ws.cell(row=row_idx, column=2).value
        if code_value is None:
            continue
        code = str(code_value).strip()
        notes: list[str] = []

        if code in INVALID_CODES:
            ws.cell(row=row_idx, column=2).fill = FILL_CODE
            notes.append(INVALID_CODES[code])
            marked_code += 1

        if code in TEAM_MISMATCH:
            ws.cell(row=row_idx, column=9).fill = FILL_TEAM
            notes.append(TEAM_MISMATCH[code])
            marked_team += 1

        if notes:
            ws.cell(row=row_idx, column=review_col, value=" / ".join(notes))

    if "검토범례" in wb.sheetnames:
        del wb["검토범례"]

    legend = wb.create_sheet("검토범례", 0)
    legend["A1"] = "색상 범례"
    legend["A1"].font = Font(bold=True, size=12)

    for col, title in zip("ABC", ["색상", "대상", "의미"], strict=True):
        cell = legend[f"{col}3"]
        cell.value = title
        cell.font = Font(bold=True)
        cell.fill = FILL_HEADER

    legend["A4"].fill = FILL_CODE
    legend["B4"] = "프로젝트코드 셀"
    legend["C4"] = "코드 형식 오류 (3건) — 웹 등록 제외"

    legend["A5"].fill = FILL_TEAM
    legend["B5"] = "담당팀 셀"
    legend["C5"] = "코드 분류와 담당팀 불일치 (6건) — 자동 대체 등록됨"

    legend["A7"] = "프로젝트코드"
    legend["B7"] = "검토 내용"
    legend["A7"].font = Font(bold=True)
    legend["B7"].font = Font(bold=True)

    row = 8
    for code, message in INVALID_CODES.items():
        legend.cell(row=row, column=1, value=code)
        legend.cell(row=row, column=2, value=message)
        row += 1

    row += 1
    legend.cell(row=row, column=1, value="프로젝트코드").font = Font(bold=True)
    legend.cell(row=row, column=2, value="담당팀 / 검토 내용").font = Font(bold=True)
    row += 1
    for code, message in TEAM_MISMATCH.items():
        legend.cell(row=row, column=1, value=code)
        legend.cell(row=row, column=2, value=message)
        row += 1

    legend.column_dimensions["A"].width = 18
    legend.column_dimensions["B"].width = 28
    legend.column_dimensions["C"].width = 48
    ws.column_dimensions[get_column_letter(review_col)].width = 56

    wb.save(OUT)
    print(f"Saved: {OUT}")
    print(f"code marked: {marked_code}, team marked: {marked_team}")


if __name__ == "__main__":
    main()
